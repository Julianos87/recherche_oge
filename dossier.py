# -*- coding: utf-8 -*-
"""
Génère un DOSSIER DE RECHERCHE COMPLET par cabinet de Géomètre-Expert, pour
faciliter la personnalisation de la lettre de motivation.

Pour chaque cabinet (regroupé par email = siège + toutes ses implantations) :
  - identité : nom, forme juridique, SIREN/SIRET, date de création ;
  - dirigeant(s) avec leur qualité ;
  - implantations (toutes les villes du cabinet) ;
  - finances : CA et résultat net par exercice disponible ;
  - effectif (tranche) ;
  - site internet (déduit du domaine email) : titre, description, spécialités
    détectées et un extrait de contenu ;
  - (option) offres de recrutement via France Travail.

Sorties :
  - sortie/dossiers/<cabinet>.md   : dossier lisible, un par cabinet
  - sortie/dossiers_cabinets.xlsx  : tableau récapitulatif
  - sortie/lettres_corrigees_V2/<region>/<departement>/lettre_<cabinet>.docx

Usage :
  python dossier.py --departements 87
  python dossier.py --region nouvelle-aquitaine --lettres
  python dossier.py --tout --pause 0.3
"""

import argparse
import re
import time
import json
import urllib.parse
import urllib.request
from collections import OrderedDict
from pathlib import Path

import openpyxl
from openpyxl.styles import Font

import pipeline as P  # réutilise les briques déjà validées

SORTIE = P.SORTIE
DOSSIERS = SORTIE / "dossiers"
LETTRES = SORTIE / "lettres_corrigees_V2"
LETTRES_SOURCE = SORTIE / "lettres_corrigees"

# Pages internes à explorer sur le site du cabinet
PAGES = ["", "/nos-prestations", "/prestations", "/nos-services", "/services",
         "/competences", "/savoir-faire", "/le-cabinet", "/qui-sommes-nous",
         "/a-propos", "/notre-cabinet"]


# --------------------------------------------------------------------------- #
# Regroupement en cabinets (siège + implantations)
# --------------------------------------------------------------------------- #
def nettoyer_nom(label):
    """Retire le suffixe d'implantation type ' - BS (VILLE)' / ' - PE (VILLE)'."""
    return re.sub(r"\s*-\s*[A-Z]{2}\s*\(.*?\)\s*$", "", label).strip()


def slugify(nom):
    return re.sub(r"[^a-z0-9]+", "-", P.sans_accent(nom).lower()).strip("-")[:60]


def region_departement(dossier):
    """Retourne les dossiers de sortie Région / Département d'un cabinet."""
    cp = str(dossier.get("zipcode", "")).strip().zfill(5)
    dep = "20" if cp.startswith("20") else cp[:2]
    region_nom = "Autre"
    for r_name, r_deps in P.REGIONS.items():
        if dep in r_deps:
            region_nom = r_name.title().replace("-", " ")
            break
    return region_nom, dep


def lire_lettre_source(dossier):
    """Lit la lettre existante correspondant au cabinet, sans modifier le fichier."""
    region_nom, dep = region_departement(dossier)
    chemin = (LETTRES_SOURCE / region_nom / dep
              / f"lettre_{slugify(dossier['nom'])}.docx")
    if not chemin.exists():
        return "", chemin

    from docx import Document
    document = Document(chemin)
    paragraphes = [p.text.strip() for p in document.paragraphs if p.text.strip()]
    return "\n".join(paragraphes), chemin


def grouper(cabinets):
    groupes = OrderedDict()
    for c in cabinets:
        cle = (c.get("email") or "").lower().strip() or c["link"]
        groupes.setdefault(cle, []).append(c)
    cabs = []
    for cle, membres in groupes.items():
        siege = next((m for m in membres if not m.get("isbureau")), membres[0])
        villes = []
        for m in membres:
            v = m.get("fullCity", "").strip()
            if v and v not in villes:
                villes.append(v)
        cabs.append({"siege": siege, "membres": membres,
                     "nom": nettoyer_nom(siege["label"]), "villes": villes})
    return cabs


# --------------------------------------------------------------------------- #
# Enrichissement approfondi (API Recherche d'entreprises)
# --------------------------------------------------------------------------- #
def enrichir_complet(siege, pause=0.3):
    cp = str(siege.get("zipcode", "")).strip()
    terme = P.patronyme_recherche(siege["label"])
    base = {"siren": "", "siret": "", "naf": "", "forme": "", "creation": "",
            "effectif": "", "dirigeants": [], "finances": {},
            "nb_etablissements": "", "nom_legal": ""}
    try:
        params = {"q": terme, "per_page": 10}
        if len(cp) == 5:
            params["code_postal"] = cp
        rep = P.api_entreprises(params)
    except Exception:
        return base
    finally:
        time.sleep(pause)

    cands = [e for e in rep.get("results", [])
             if (e.get("activite_principale") or "").startswith("71.12")]
    if not cands:
        return base
    e = max(cands, key=lambda x: P.similarite(x.get("nom_complet", ""), siege["label"]))

    base["nom_legal"] = e.get("nom_complet", "")
    base["siren"] = e.get("siren", "")
    base["siret"] = (e.get("siege") or {}).get("siret", "")
    base["naf"] = e.get("activite_principale", "")
    base["forme"] = e.get("nature_juridique", "")
    base["creation"] = (e.get("date_creation") or "")[:4]
    base["effectif"] = P.libelle_effectif(e.get("tranche_effectif_salarie"))
    base["nb_etablissements"] = e.get("nombre_etablissements", "")
    for d in e.get("dirigeants", []):
        nom = d.get("denomination") or " ".join(
            x for x in [(d.get("prenoms") or "").title(),
                        (d.get("nom") or "").title()] if x).strip()
        if nom:
            base["dirigeants"].append({"nom": nom, "qualite": d.get("qualite", "")})
    base["finances"] = e.get("finances") or {}
    return base


def noms_dirigeants(ent):
    return " ; ".join(d["nom"] for d in ent.get("dirigeants", [])[:4]) or "Non disponible"


# --------------------------------------------------------------------------- #
# Scraping approfondi du site propre du cabinet
# --------------------------------------------------------------------------- #
def nettoyer_html(h):
    import html as _html
    # Supprimer les balises inutiles (scripts, navigation, header/footer)
    h = re.sub(r"(?is)<(script|style|noscript|nav|header|footer).*?</\1>", " ", h)
    # Remplacer les balises de blocs par des sauts de ligne
    h = re.sub(r"(?is)</?(p|div|h[1-6]|li|br|section|article)[^>]*>", "\n", h)
    # Supprimer toutes les autres balises
    h = re.sub(r"(?is)<[^>]+>", " ", h)
    h = _html.unescape(h)
    
    # Ne garder que les lignes qui ressemblent à du vrai texte (phrases complètes)
    lignes_propres = []
    for ligne in h.split("\n"):
        ligne = re.sub(r"\s+", " ", ligne).strip()
        # Garder si + de 40 caractères ou + de 5 mots
        if len(ligne) > 40 or len(ligne.split()) > 5:
            lignes_propres.append(ligne)
            
    return "\n".join(lignes_propres)


def scraper_site(cabinet_siege, pages_max=6):
    site = P.decouvrir_site(cabinet_siege)
    res = {"url": site or "", "titre": "", "description": "",
           "specialites": "Non disponible", "extrait": ""}
    if not site:
        return res
    textes, brut_total = [], ""
    site_actif = site
    
    # Test d'accessibilité de la page d'accueil avec fallback sur www.
    try:
        h_accueil = P.http_get(site_actif, timeout=12).decode("utf-8", "replace")
    except Exception:
        if "://" in site_actif and "://www." not in site_actif:
            site_actif = site_actif.replace("://", "://www.")
            try:
                h_accueil = P.http_get(site_actif, timeout=12).decode("utf-8", "replace")
            except Exception:
                return res
        else:
            return res
            
    res["url"] = site_actif
    
    for chemin in PAGES[:pages_max]:
        try:
            if chemin == "":
                h = h_accueil
            else:
                h = P.http_get(site_actif + chemin, timeout=12).decode("utf-8", "replace")
        except Exception:
            continue
        brut_total += " " + h
        if not res["titre"]:
            m = re.search(r"(?is)<title>(.*?)</title>", h)
            if m:
                res["titre"] = nettoyer_html(m.group(1))[:160]
        if not res["description"]:
            m = re.search(r'(?is)<meta[^>]+name=["\']description["\'][^>]+'
                          r'content=["\'](.*?)["\']', h)
            if m:
                res["description"] = nettoyer_html(m.group(1))[:300]
        textes.append(nettoyer_html(h))
    if not brut_total.strip():
        return res
    bas = P.sans_accent(brut_total).lower()
    found = [m for m in P.MOTS_SPECIALITES if P.sans_accent(m).lower() in bas]
    if found:
        res["specialites"] = ", ".join(found)
    # extrait : déduplication de tous les paragraphes propres récoltés
    toutes_les_lignes = []
    for t in textes:
        toutes_les_lignes.extend(t.split("\n"))
        
    lignes_uniques = []
    vus = set()
    for ligne in toutes_les_lignes:
        if ligne and ligne not in vus:
            vus.add(ligne)
            lignes_uniques.append(ligne)
            
    res["extrait"] = "\n".join(lignes_uniques)[:10000]
    if len("\n".join(lignes_uniques)) > 10000:
        res["extrait"] += "…"
    return res


# --------------------------------------------------------------------------- #
# Construction et rendu du dossier
# --------------------------------------------------------------------------- #
def construire(cab, ent, web, recrutement, analyse_ia=""):
    s = cab["siege"]
    return {
        "nom": cab["nom"],
        "nom_legal": ent.get("nom_legal", ""),
        "dirigeants": ent.get("dirigeants", []),
        "dirigeant_principal": (ent["dirigeants"][0]["nom"]
                                if ent.get("dirigeants") else "Non disponible"),
        "adresse": ", ".join(x for x in [s.get("fullAddress", "").strip(),
                                         s.get("fullCity", "").strip()] if x),
        "zipcode": str(s.get("zipcode", "")).strip(),
        "villes": cab["villes"],
        "email": s.get("email", ""), "tel": s.get("phone", ""),
        "site": web.get("url", ""), "site_titre": web.get("titre", ""),
        "site_description": web.get("description", ""),
        "site_extrait": web.get("extrait", ""),
        "specialites": web.get("specialites", "Non disponible"),
        "forme": ent.get("forme", ""), "siren": ent.get("siren", ""),
        "siret": ent.get("siret", ""), "creation": ent.get("creation", ""),
        "effectif": ent.get("effectif", ""),
        "nb_etablissements": ent.get("nb_etablissements", ""),
        "finances": ent.get("finances", {}),
        "recrutement": recrutement,
        "fiche_oge": s.get("link", ""),
        "analyse_ia": analyse_ia,
    }


def finances_lignes(fin):
    out = []
    for annee in sorted(fin.keys()):
        d = fin[annee]
        ca = d.get("ca")
        rn = d.get("resultat_net")
        bout = f"{annee} : "
        bout += (f"CA {int(ca):,} €".replace(",", " ") if ca else "CA n.c.")
        if rn is not None:
            bout += f", résultat net {int(rn):,} €".replace(",", " ")
        out.append(bout)
    return out


def rendre_md(d):
    L = [f"# {d['nom']}", ""]
    if d["nom_legal"] and d["nom_legal"].upper() != d["nom"].upper():
        L.append(f"*Raison sociale : {d['nom_legal']}*\n")
    L += ["## Contact", f"- Adresse (siège) : {d['adresse'] or 'Non disponible'}"]
    if len(d["villes"]) > 1:
        L.append(f"- Implantations ({len(d['villes'])}) : " + " · ".join(d["villes"]))
    L.append(f"- Téléphone : {d['tel'] or 'Non disponible'}")
    L.append(f"- Email : {d['email'] or 'Non disponible'}")
    L.append(f"- Site internet : {d['site'] or 'Non disponible'}")
    L.append(f"- Fiche OGE : {d['fiche_oge']}")

    L += ["", "## Dirigeant(s) (Géomètre(s)-Expert(s))"]
    if d["dirigeants"]:
        for dd in d["dirigeants"]:
            q = f" — {dd['qualite']}" if dd.get("qualite") else ""
            L.append(f"- {dd['nom']}{q}")
    else:
        L.append("- Non disponible")

    L += ["", "## Identité & structure",
          f"- Forme juridique : {d['forme'] or 'Non disponible'}",
          f"- SIREN : {d['siren'] or 'Non disponible'}"
          f"  |  SIRET (siège) : {d['siret'] or 'Non disponible'}",
          f"- Création : {d['creation'] or 'Non disponible'}",
          f"- Effectif : {d['effectif'] or 'Non disponible'}",
          f"- Nombre d'établissements : {d['nb_etablissements'] or 'Non disponible'}"]

    L += ["", "## Finances (source : API entreprises / comptes publiés)"]
    fl = finances_lignes(d["finances"])
    L += [f"- {x}" for x in fl] if fl else ["- Non disponible (comptes non publiés)"]

    L += ["", "## Spécialités & activité",
          f"- Spécialités détectées : {d['specialites']}"]
    if d["site_titre"]:
        L.append(f"- Titre du site : {d['site_titre']}")
    if d["site_description"]:
        L.append(f"- Description : {d['site_description']}")

    L += ["", "## Recrutement", f"- {d['recrutement']}"]

    if d.get("analyse_ia"):
        L += ["", "## Analyse intelligente du profil du cabinet", d["analyse_ia"]]

    if d["site_extrait"]:
        L += ["", "## Extrait du site (brut)", "> " + d["site_extrait"].replace("\n", " ")]
    return "\n".join(L) + "\n"


def dossier_texte_pour_llm(d):
    """Version condensée injectée dans le prompt de génération de lettre."""
    def disponible(valeur):
        texte = str(valeur or "").strip().lower()
        return texte not in {"", "non disponible", "n.c.", "none"}

    donnees_site = any([
        disponible(d.get("site_description")),
        disponible(d.get("site_extrait")),
        disponible(d.get("analyse_ia")),
        disponible(d.get("specialites")),
    ])
    donnees_structure = any([
        len(d.get("villes") or []) > 1,
        disponible(d.get("creation")),
        disponible(d.get("recrutement")),
    ])

    if donnees_site:
        niveau = "RICHE"
        consigne = "Conserver cinq paragraphes argumentatifs : vision du métier, VOUS précis, atouts, puis immersion et apprentissage."
    elif donnees_structure:
        niveau = "LIMITÉ"
        consigne = "Conserver quatre ou cinq paragraphes argumentatifs : aucun VOUS court, mais fondre les rares faits avec les atouts et la contribution."
    else:
        niveau = "MINIMAL"
        consigne = "Conserver quatre paragraphes argumentatifs sans VOUS autonome : projet, vision du métier, atouts, puis immersion et apprentissage."

    parts = [
        f"NIVEAU DE PERSONNALISATION : {niveau}",
        f"CONSIGNE : {consigne}",
        f"Nom : {d['nom']}",
        f"Dirigeant(s) : {d['dirigeant_principal']}",
        f"Adresse siège : {d['adresse'] or 'Non disponible'}",
    ]
    if len(d["villes"]) > 1:
        parts.append("Implantations : " + ", ".join(d["villes"]))
    if d["creation"]:
        parts.append(f"Historique vérifié : créé en {d['creation']}")
    if disponible(d.get("specialites")):
        parts.append(f"Spécialités documentées : {d['specialites']}")
    if d["site_description"]:
        parts.append(f"Présentation (site) : {d['site_description']}")
    if d.get("site_extrait"):
        parts.append(f"Détails extraits du site : {d['site_extrait']}")
    if d.get("analyse_ia"):
        parts.append(f"Analyse IA du site : {d['analyse_ia']}")
    if disponible(d.get("recrutement")):
        parts.append(f"Recrutement : {d['recrutement']}")
    return "\n".join(parts)

# --------------------------------------------------------------------------- #
# Génération de lettre (modèle de style + dossier)
# --------------------------------------------------------------------------- #
def appel_llm_local(prompt):
    """Génère via un serveur LOCAL compatible OpenAI (LM Studio, llama.cpp, vLLM…).

    Configuration par variables d'environnement :
      LOCAL_LLM_URL   (défaut http://localhost:1234/v1/chat/completions  — LM Studio)
      LOCAL_LLM_MODEL (défaut "local-model" ; mets l'id du modèle chargé si besoin)
    """
    import os
    url = os.environ.get("LOCAL_LLM_URL", "http://localhost:1234/v1/chat/completions")
    modele = os.environ.get("LOCAL_LLM_MODEL", "local-model")
    corps = json.dumps({
        "model": modele,
        "messages": [{"role": "user", "content": prompt}],
        "temperature": 0.2,
        "max_tokens": -1,  # -1 signifie AUCUNE limite de taille
        "stream": False,
    }).encode("utf-8")
    req = urllib.request.Request(url, data=corps,
                                 headers={"Content-Type": "application/json"})
    # Timeout augmenté à 1200 secondes (20 minutes) pour laisser le temps à LM Studio
    with urllib.request.urlopen(req, timeout=1200) as r:
        rep = json.loads(r.read())
    texte = rep["choices"][0]["message"]["content"].strip()
    # Nettoyage des balises markdown (ex: ```markdown ... ```)
    if texte.startswith("```"):
        texte = re.sub(r"^```[a-zA-Z]*\n", "", texte)
        texte = re.sub(r"\n```$", "", texte)
    return texte.strip()


def appel_llm_anthropic(prompt):
    import os
    try:
        import anthropic
    except ImportError:
        return None
    cle = os.environ.get("ANTHROPIC_API_KEY")
    if not cle:
        return None
    client = anthropic.Anthropic(api_key=cle)
    msg = client.messages.create(model="claude-opus-4-8", max_tokens=1800,
                                 messages=[{"role": "user", "content": prompt}])
    return "".join(b.text for b in msg.content if getattr(b, "type", "") == "text")


def generer_lettre(d, profil, lettre_modele, gabarit, lettre_existante,
                   backend="anthropic"):
    prompt = gabarit.format(profil_candidat=profil, lettre_modele=lettre_modele,
                            dossier_cabinet=dossier_texte_pour_llm(d),
                            lettre_existante=lettre_existante)
    try:
        if backend == "local":
            return appel_llm_local(prompt)
        return appel_llm_anthropic(prompt)
    except Exception as e:
        print(f"    [lettre] échec génération ({backend}) : {e}")
        return None


# --------------------------------------------------------------------------- #
# Export récapitulatif
# --------------------------------------------------------------------------- #
COLS = [("nom", "Cabinet"), ("dirigeant_principal", "Dirigeant"),
        ("adresse", "Siège"), ("_implantations", "Implantations"),
        ("_finances", "Finances"), ("effectif", "Effectif"),
        ("specialites", "Spécialités"), ("recrutement", "Recrutement"),
        ("site", "Site"), ("email", "Email"), ("tel", "Tél"), ("siren", "SIREN")]


def exporter_xlsx(dossiers, chemin):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cabinets"
    for j, (_, t) in enumerate(COLS, 1):
        ws.cell(1, j, t).font = Font(bold=True)
    for i, d in enumerate(dossiers, 2):
        d = dict(d)
        d["_implantations"] = " · ".join(d["villes"])
        d["_finances"] = " ; ".join(finances_lignes(d["finances"])) or ""
        for j, (k, _) in enumerate(COLS, 1):
            ws.cell(i, j, d.get(k, ""))
    larg = [30, 24, 38, 30, 40, 16, 34, 30, 28, 26, 14, 12]
    for j, w in enumerate(larg, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(j)].width = w
    ws.freeze_panes = "A2"
    wb.save(chemin)


# --------------------------------------------------------------------------- #
# Main
# --------------------------------------------------------------------------- #
def main():
    ap = argparse.ArgumentParser(description="Dossiers de recherche par cabinet GE.")
    ap.add_argument("--departements", nargs="+")
    ap.add_argument("--region", choices=sorted(P.REGIONS))
    ap.add_argument("--tout", action="store_true")
    ap.add_argument("--lettres", action="store_true",
                    help="générer les lettres .docx")
    ap.add_argument("--local", action="store_true",
                    help="générer les lettres via un LLM LOCAL (LM Studio / OpenAI-compatible)")
    ap.add_argument("--lettres-seules", dest="lettres_seules", action="store_true",
                    help="ne génère QUE les lettres depuis sortie/dossiers.json "
                         "(aucun appel API/scraping ; idéal en local GPU)")
    ap.add_argument("--analyse-site", action="store_true",
                    help="analyse l'extrait du site avec l'IA pour enrichir le dossier")
    ap.add_argument("--recrutement", action="store_true",
                    help="détecter les offres France Travail (FT_CLIENT_ID/SECRET)")
    ap.add_argument("--pause", type=float, default=0.35)
    ap.add_argument("--max", type=int, default=0)
    args = ap.parse_args()

    SORTIE.mkdir(exist_ok=True)
    DOSSIERS.mkdir(exist_ok=True)
    LETTRES.mkdir(exist_ok=True)
    
    # Pré-création de toute l'arborescence (Régions -> Départements)
    for r_name, r_deps in P.REGIONS.items():
        region_nom = r_name.title().replace("-", " ")
        for dep in r_deps:
            (DOSSIERS / region_nom / dep).mkdir(parents=True, exist_ok=True)
            (LETTRES / region_nom / dep).mkdir(parents=True, exist_ok=True)

    racine = Path(__file__).resolve().parent
    profil = (racine / "profil_candidat.txt").read_text(encoding="utf-8")
    gabarit = (racine / "prompt_lettre.txt").read_text(encoding="utf-8")
    modele = SORTIE / "_lettre_julian.txt"
    lettre_modele = modele.read_text(encoding="utf-8") if modele.exists() else ""
    chemin_json = SORTIE / "dossiers.json"

    backend = "local" if args.local else "anthropic"

    # ----- Mode "lettres seules" : pas de réseau, on lit le cache JSON -------
    if args.lettres_seules:
        if not chemin_json.exists():
            ap.error(f"{chemin_json} introuvable : lance d'abord une collecte.")
        dossiers = json.loads(chemin_json.read_text(encoding="utf-8"))
        if args.region:
            deps_autorises = set(P.REGIONS[args.region])
            dossiers = [d for d in dossiers
                        if region_departement(d)[1] in deps_autorises]
        elif args.departements:
            deps_autorises = {d.zfill(2) for d in args.departements}
            dossiers = [d for d in dossiers
                        if region_departement(d)[1] in deps_autorises]
        if args.max:
            dossiers = dossiers[:args.max]
        print(f"{len(dossiers)} dossiers -> génération des lettres ({backend})...")
        for n, d in enumerate(dossiers, 1):
            print(f"  [{n}/{len(dossiers)}] {d['nom']}")
            lettre_existante, chemin_source = lire_lettre_source(d)
            if not lettre_existante:
                print(f"    [lettre] source introuvable, ignorée : {chemin_source}")
                continue
            txt = generer_lettre(d, profil, lettre_modele, gabarit,
                                 lettre_existante, backend)
            if txt:
                region_nom, dep = region_departement(d)
                lettre_region_dep = LETTRES / region_nom / dep
                lettre_region_dep.mkdir(parents=True, exist_ok=True)
                chemin_lettre = lettre_region_dep / f"lettre_{slugify(d['nom'])}.docx"
                P.ecrire_docx(txt, chemin_lettre)
                if not chemin_lettre.exists():
                    raise OSError(f"Lettre non créée : {chemin_lettre}")
                print(f"    -> enregistrée : {chemin_lettre.resolve()}")
        print(f"\nTerminé. Lettres V2 : {LETTRES}")
        return

    # ----- Mode collecte ----------------------------------------------------
    deps = None
    if args.region:
        deps = P.REGIONS[args.region]
    elif args.departements:
        deps = [d.zfill(2) for d in args.departements]
    elif not args.tout:
        ap.error("Précise --departements, --region, --tout ou --lettres-seules.")

    # on garde tous les membres (sièges + bureaux) pour reconstituer les implantations
    bruts = P.filtrer(P.charger_cabinets(), deps, sieges_seuls=False)
    cabs = grouper(bruts)
    if args.max:
        cabs = cabs[:args.max]
    print(f"{len(cabs)} cabinets (regroupés) à traiter.")

    token_ft = P.ft_token() if args.recrutement else None
    if args.recrutement and not token_ft:
        print("  [France Travail] identifiants absents -> recrutement ignoré.")

    dossiers = []
    deja_traites = set()
    if chemin_json.exists():
        try:
            anciens = json.loads(chemin_json.read_text(encoding="utf-8"))
            anciens_valides = []
            for a in anciens:
                slug = slugify(a["nom"])
                # Retrouver le chemin du .md
                cp = str(a.get("zipcode", "")).strip().zfill(5)
                dep = cp[:2]
                if cp.startswith("20"): dep = "20"
                region_nom = "Autre"
                for r_name, r_deps in P.REGIONS.items():
                    if dep in r_deps:
                        region_nom = r_name.title().replace("-", " ")
                        break
                chemin_md = DOSSIERS / region_nom / dep / f"{slug}.md"
                
                # Si le fichier .md a été supprimé par l'utilisateur, on l'ignore du cache
                if chemin_md.exists():
                    anciens_valides.append(a)
                    deja_traites.add(slug)
                    
            dossiers.extend(anciens_valides)
            if deja_traites:
                print(f"Reprise activée : {len(deja_traites)} cabinets intacts trouvés dans le cache.")
        except Exception:
            pass

    for n, cab in enumerate(cabs, 1):
        slug = slugify(cab["nom"])
        if slug in deja_traites:
            print(f"  [{n}/{len(cabs)}] {cab['nom']} -> Déjà traité (ignoré)")
            continue
            
        print(f"  [{n}/{len(cabs)}] {cab['nom']}")
        ent = enrichir_complet(cab["siege"], pause=args.pause)
        web = scraper_site(cab["siege"])
        recr = (P.offres_recrutement(cab["siege"], token_ft)
                if token_ft else "Non disponible")
                
        analyse_ia = ""
        if args.analyse_site and web.get("extrait"):
            prompt_analyse = (
                f"Voici le texte extrait du site web d'un cabinet de géomètres-experts :\n"
                f'"{web["extrait"]}"\n\n'
                f"Fais une analyse TRÈS DÉTAILLÉE et exhaustive en 3 parties pour m'aider à hyper-personnaliser ma lettre de motivation. Je veux un maximum d'informations tirées du texte :\n"
                f"1. Activités & Expertises clés (liste en détail les prestations, spécialités, ou technologies mentionnées)\n"
                f"2. Valeurs, Histoire & Philosophie du cabinet (engagements, historique, approche client)\n"
                f"3. Mots-clés lexicaux et ambiance (le vocabulaire spécifique qu'ils utilisent, l'esprit d'équipe)\n"
                f"Ne renvoie QUE ton analyse détaillée, sans introduction ni conclusion."
            )
            try:
                print("    -> Analyse IA du site en cours...")
                if backend == "local":
                    analyse_ia = appel_llm_local(prompt_analyse)
                else:
                    analyse_ia = appel_llm_anthropic(prompt_analyse)
            except Exception as e:
                print(f"    [analyse site] échec : {e}")
                analyse_ia = "Erreur lors de l'analyse IA."
                
        d = construire(cab, ent, web, recr, analyse_ia)
        dossiers.append(d)

        # Détermination de la région et du département
        region_nom, dep = region_departement(d)
                
        # Création des sous-dossiers pour les dossiers MD
        dossier_region_dep = DOSSIERS / region_nom / dep
        dossier_region_dep.mkdir(parents=True, exist_ok=True)
        chemin_md = dossier_region_dep / f"{slug}.md"
        chemin_md.write_text(rendre_md(d), encoding="utf-8")
        
        # sauvegarde incrémentale du cache JSON (résiste à une interruption)
        chemin_json.write_text(json.dumps(dossiers, ensure_ascii=False, indent=1),
                               encoding="utf-8")

        if args.lettres and lettre_modele:
            lettre_region_dep = LETTRES / region_nom / dep
            lettre_region_dep.mkdir(parents=True, exist_ok=True)
            lettre_existante, chemin_source = lire_lettre_source(d)
            if not lettre_existante:
                print(f"    [lettre] source introuvable, ignorée : {chemin_source}")
            else:
                txt = generer_lettre(d, profil, lettre_modele, gabarit,
                                     lettre_existante, backend)
                if txt:
                    chemin_lettre = lettre_region_dep / f"lettre_{slug}.docx"
                    P.ecrire_docx(txt, chemin_lettre)
                    if not chemin_lettre.exists():
                        raise OSError(f"Lettre non créée : {chemin_lettre}")
                    print(f"    -> enregistrée : {chemin_lettre.resolve()}")

    exporter_xlsx(dossiers, SORTIE / "dossiers_cabinets.xlsx")
    print(f"\nTerminé. Dossiers : {DOSSIERS}\nRécapitulatif : {SORTIE/'dossiers_cabinets.xlsx'}"
          f"\nCache : {chemin_json}")


if __name__ == "__main__":
    main()
